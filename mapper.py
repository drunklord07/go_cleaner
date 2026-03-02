#!/usr/bin/env python3
"""
GoPhish -> Userbase Identity Mapper + CSAM Phishing Status
Combines GoPhish campaign CSVs, enriches with userbase data,
then cross-references CSAM submit/clicked logs.
"""

import os, sys, glob
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# USERBASE COLUMN CONFIGURATION
# Check these match your Userbase Excel headers exactly (case-sensitive).
# If a column has a different name in your file, update ONLY these lines.
# ==============================================================================

UB_COL_EMPLOYEE_NAME  = "Employee Name"                    # Full name column
UB_COL_EMPLOYEE_EMAIL = "Employee email"                   # Email column
UB_COL_SAVIYNT_SSOUPN = "SSOUPN as per Saviynt"           # Saviynt SSO ID column
UB_COL_AD_SSOUPN      = "locationdesc (SSOUPN as per AD)" # AD SSO ID column

# ==============================================================================
# CSAM FILE NAMES  (must sit in same folder as the Userbase file)
# ==============================================================================

CSAM_SUBMIT_FILENAME  = "csam submit logs.xlsx"           # CSAM submitted data file
CSAM_CLICKED_FILENAME = "csam - clicked logs.xlsx"        # CSAM clicked link file

# ==============================================================================
# CSAM COLUMN NAMES
# ==============================================================================

CSAM_COL_EMAIL_UPN = "Email/UPN"                          # Email/UPN col in both CSAM files
CSAM_COL_USERNAME  = "Username that was fileed by user"   # Username col in submit file only

# ==============================================================================


# -- Colours -------------------------------------------------------------------
C = {
    "nav_dark":   "1F3864", "nav_mid":   "2E4057", "white":     "FFFFFF",
    "gp_hdr":     "808080", "row_alt":   "F5F5F5",
    "map_100":    "C6EFCE", "map_75":    "FFEB9C", "map_50":    "DDEEFF",
    "map_none":   "FCE4EC",
    "csam_sub":   "D5E8D4", "csam_clk":  "FFF2CC", "csam_none": "F8CECC",
    "ok":         "C6EFCE", "err":       "FFC7CE", "warn":      "FFEB9C",
}
THIN = Side(style="thin", color="CCCCCC")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def mkfill(c):
    return PatternFill("solid", fgColor=c)

def mkfont(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def mkalign(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def ts():
    return datetime.now().strftime("%H:%M:%S")

def log(m, i=0):   print(f"[{ts()}]{'  '*i} {m}")
def ok_(m, i=1):   print(f"[{ts()}]{'  '*i} OK  {m}")
def err_(m, i=1):  print(f"[{ts()}]{'  '*i} ERR {m}")
def info_(m, i=1): print(f"[{ts()}]{'  '*i} >>  {m}")


# -- Name / email helpers ------------------------------------------------------
def split_name(full):
    if not full or str(full).strip() in ("", "nan"):
        return "", ""
    p = str(full).strip().split()
    return (p[0], " ".join(p[1:])) if len(p) > 1 else (p[0], "")

def extract_local(email):
    if not email or str(email).strip() in ("", "nan"):
        return ""
    return str(email).strip().lower().split("@")[0]

def local_to_names(local):
    if "." in local:
        p = local.split(".", 1)
        return p[0].strip(), p[1].strip()
    return "", ""

def norm(s):
    return str(s).strip().lower() if s and str(s).strip() not in ("", "nan") else ""


# -- Status sheet --------------------------------------------------------------
class StatusSheet:
    def __init__(self, wb):
        self.ws    = wb.create_sheet("Status", 0)
        self.steps = []
        self.nrow  = 6
        self._setup()

    def _setup(self):
        ws = self.ws
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:H1")
        cell = ws["A1"]
        cell.value     = "GoPhish  Userbase Mapper + CSAM  |  Run Checklist"
        cell.font      = mkfont(bold=True, color=C["white"], size=13)
        cell.fill      = mkfill(C["nav_dark"])
        cell.alignment = mkalign("center", "center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:H2")
        cell = ws["A2"]
        cell.value     = f"Run: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
        cell.font      = mkfont(italic=True, size=9, color="555555")
        cell.alignment = mkalign("center")
        ws.row_dimensions[2].height = 15

        ws.merge_cells("A3:H3")
        cell = ws["A3"]
        cell.value = "   OK = Done        ERR = Error        >> = Running        WARN = Skipped/Warning"
        cell.font      = mkfont(size=9, color="333333")
        cell.fill      = mkfill("F0F0F0")
        cell.alignment = mkalign("left", "center")
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 6

        hdrs   = ["#", "Step Description", "Status", "Processed", "Matched", "Errors", "Time", "Notes"]
        widths = [4,    44,                  12,       11,          10,        8,         8,      54]
        for col, (h, w) in enumerate(zip(hdrs, widths), 1):
            cell = ws.cell(row=5, column=col)
            cell.value     = h
            cell.font      = mkfont(bold=True, color=C["white"], size=10)
            cell.fill      = mkfill(C["nav_mid"])
            cell.alignment = mkalign("center", "center", wrap=True)
            cell.border    = BORD
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[5].height = 24
        ws.freeze_panes = "A6"

    def _bg(self, s):
        if "OK"   in s: return C["ok"]
        if "ERR"  in s: return C["err"]
        if "WARN" in s: return C["warn"]
        return "EEF2FF"

    def add(self, desc, status=">> Running", notes=""):
        ws, row = self.ws, self.nrow
        for col, val in enumerate([len(self.steps)+1, desc, status, "", "", "", "", notes], 1):
            cell = ws.cell(row=row, column=col)
            cell.value     = val
            cell.font      = mkfont(size=10, bold=(col == 2))
            cell.fill      = mkfill(self._bg(status))
            cell.alignment = mkalign("left" if col in (2, 8) else "center",
                                     "center", wrap=(col in (2, 8)))
            cell.border    = BORD
        ws.row_dimensions[row].height = 20
        self.steps.append({"row": row, "start": datetime.now()})
        self.nrow += 1
        return len(self.steps) - 1

    def update(self, idx, status, processed=None, matched=None, errors=None, notes=None):
        ws, row = self.ws, self.steps[idx]["row"]
        elapsed = (datetime.now() - self.steps[idx]["start"]).total_seconds()
        bg = self._bg(status)
        updates = {3: status, 7: f"{elapsed:.1f}s"}
        if processed is not None: updates[4] = processed
        if matched   is not None: updates[5] = matched
        if errors    is not None: updates[6] = errors
        if notes     is not None: updates[8] = notes
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if col in updates: cell.value = updates[col]
            cell.fill      = mkfill(bg)
            cell.font      = mkfont(size=10, bold=(col == 2))
            cell.alignment = mkalign("left" if col in (2, 8) else "center",
                                     "center", wrap=(col in (2, 8)))


# -- Sheet helpers -------------------------------------------------------------
def write_header_row(ws, headers, row=1, bg=C["nav_dark"], fg=C["white"], height=22):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value     = h
        cell.font      = mkfont(bold=True, color=fg, size=10)
        cell.fill      = mkfill(bg)
        cell.alignment = mkalign("center", "center", wrap=True)
        cell.border    = BORD
    ws.row_dimensions[row].height = height

def auto_width(ws, mn=8, mx=46):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[letter].width = min(max(w + 2, mn), mx)

def build_index(series, df):
    idx = {}
    for i, val in series.items():
        k = norm(val)
        if k and k not in idx:
            idx[k] = df.loc[i].to_dict()
    return idx

def find_col(df, exact, label=""):
    if exact in df.columns:
        return exact
    matches = [c for c in df.columns if c.lower() == exact.lower()]
    if matches:
        err_(f"Column '{exact}' not found ({label}) -- did you mean '{matches[0]}'? "
             f"Update the config block at the top of the script.")
        return matches[0]
    err_(f"Column '{exact}' NOT FOUND ({label}). Available: {list(df.columns)}")
    return None

def progress_bar(i, total, start, extra=""):
    pct    = (i + 1) / total
    filled = int(40 * pct)
    bar    = "#" * filled + "." * (40 - filled)
    elapsed = (datetime.now() - start).total_seconds()
    rate   = (i + 1) / elapsed if elapsed else 0
    eta    = (total - i - 1) / rate if rate else 0
    eta_s  = f"{int(eta//60)}m{int(eta%60):02d}s" if eta > 60 else f"{eta:.0f}s"
    print(f"\r  [{bar}] {pct*100:5.1f}%  {i+1}/{total}  {extra}  {rate:.0f}/s  ETA:{eta_s}   ",
          end="", flush=True)


# -- Main ----------------------------------------------------------------------
def main():
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    input_dir   = os.path.join(script_dir, "input")
    output_path = os.path.join(script_dir, "gophish_mapped.xlsx")

    # Userbase = first xlsx in script_dir that is NOT the output or CSAM files
    skip = {"gophish_mapped.xlsx",
            CSAM_SUBMIT_FILENAME.lower(),
            CSAM_CLICKED_FILENAME.lower()}
    ub_candidates = [f for f in glob.glob(os.path.join(script_dir, "*.xlsx"))
                     if os.path.basename(f).lower() not in skip]
    userbase_path = ub_candidates[0] if ub_candidates else None
    base_dir      = os.path.dirname(userbase_path) if userbase_path else script_dir

    print("=" * 68)
    log("GoPhish -> Userbase Mapper + CSAM  |  starting")
    log(f"Input dir   : {input_dir}")
    log(f"Userbase    : {userbase_path or 'NOT FOUND'}")
    log(f"CSAM Submit : {os.path.join(base_dir, CSAM_SUBMIT_FILENAME)}")
    log(f"CSAM Clicked: {os.path.join(base_dir, CSAM_CLICKED_FILENAME)}")
    log(f"Output      : {output_path}")
    print("=" * 68)

    wb = Workbook()
    wb.active.title = "_tmp"
    st = StatusSheet(wb)

    # ---- STEP 1: Find CSVs --------------------------------------------------
    log("\n[STEP 1] Scanning input/ for GoPhish CSVs ...")
    s1   = st.add("Scan input/ for GoPhish CSV files")
    csvs = sorted(glob.glob(os.path.join(input_dir, "*.csv")))
    if not csvs:
        err_(f"No CSVs in {input_dir}")
        st.update(s1, "ERR", notes=f"No CSVs in {input_dir}")
        wb.save(output_path)
        sys.exit(1)
    ok_(f"Found {len(csvs)} file(s): {', '.join(os.path.basename(f) for f in csvs)}")
    st.update(s1, "OK", processed=len(csvs),
              notes=", ".join(os.path.basename(f) for f in csvs))

    # ---- STEP 2: Combine CSVs -----------------------------------------------
    log("\n[STEP 2] Combining CSVs ...")
    s2 = st.add("Combine all GoPhish CSVs into one DataFrame")
    try:
        frames = []
        for fp in csvs:
            df_tmp = pd.read_csv(fp, dtype=str, encoding="utf-8-sig")
            df_tmp.columns = [c.strip() for c in df_tmp.columns]
            df_tmp["Campaign Name"] = os.path.splitext(os.path.basename(fp))[0]
            frames.append(df_tmp)
        df_gp = pd.concat(frames, ignore_index=True)

        col_map = {}
        for c in df_gp.columns:
            cl = c.lower().replace(" ", "_")
            if cl == "campaign_name":                  col_map[c] = "Campaign Name"
            elif "campaign" in cl and "id" in cl:      col_map[c] = "campaign_id"
            elif cl == "email":                        col_map[c] = "email"
            elif "time" in cl:                         col_map[c] = "time"
            elif "message" in cl:                      col_map[c] = "message"
            elif "detail" in cl:                       col_map[c] = "details"
        df_gp.rename(columns=col_map, inplace=True)
        for req in ["campaign_id", "email", "time", "message", "details", "Campaign Name"]:
            if req not in df_gp.columns:
                df_gp[req] = ""

        ok_(f"Combined {len(df_gp)} rows from {len(csvs)} files")
        st.update(s2, "OK", processed=len(df_gp))
    except Exception as e:
        err_(str(e))
        st.update(s2, "ERR", notes=str(e))
        wb.save(output_path)
        sys.exit(1)

    # ---- STEP 3: Split GoPhish email -> name --------------------------------
    log("\n[STEP 3] Splitting GoPhish email -> First / Last name ...")
    s3 = st.add("Split GoPhish email into GP_First_Name / GP_Last_Name")
    df_gp["_local"]        = df_gp["email"].apply(extract_local)
    df_gp["GP_First_Name"] = df_gp["_local"].apply(lambda x: local_to_names(x)[0])
    df_gp["GP_Last_Name"]  = df_gp["_local"].apply(lambda x: local_to_names(x)[1])
    named = (df_gp["GP_First_Name"] != "").sum()
    ok_(f"Name extracted for {named}/{len(df_gp)} rows (first.last format)")
    st.update(s3, "OK", processed=len(df_gp), matched=named,
              notes=f"{len(df_gp)-named} rows have ID-format emails")

    # ---- STEP 4: Load Userbase ----------------------------------------------
    log("\n[STEP 4] Loading Userbase Excel ...")
    s4 = st.add("Load Userbase Excel file")
    if not userbase_path:
        err_("Userbase .xlsx not found beside the script!")
        st.update(s4, "ERR", notes="No .xlsx found in script directory")
        wb.save(output_path)
        sys.exit(1)
    try:
        df_ub = pd.read_excel(userbase_path, dtype=str)
        df_ub.columns = [c.strip() for c in df_ub.columns]
        ok_(f"Loaded {len(df_ub)} rows | cols: {list(df_ub.columns)}")
        st.update(s4, "OK", processed=len(df_ub))
    except Exception as e:
        err_(str(e))
        st.update(s4, "ERR", notes=str(e))
        wb.save(output_path)
        sys.exit(1)

    # ---- STEP 5: Enrich Userbase --------------------------------------------
    log("\n[STEP 5] Enriching Userbase ...")
    s5 = st.add("Enrich Userbase: split names, extract locals from emails/SSO IDs")

    ub_name_col  = find_col(df_ub, UB_COL_EMPLOYEE_NAME,  "Employee Name")
    ub_email_col = find_col(df_ub, UB_COL_EMPLOYEE_EMAIL, "Employee Email")
    ub_sav_col   = find_col(df_ub, UB_COL_SAVIYNT_SSOUPN, "Saviynt SSOUPN")
    ub_ad_col    = find_col(df_ub, UB_COL_AD_SSOUPN,      "AD SSOUPN")

    info_(f"Name col    : {ub_name_col}")
    info_(f"Email col   : {ub_email_col}")
    info_(f"Saviynt col : {ub_sav_col}")
    info_(f"AD col      : {ub_ad_col}")

    if ub_name_col:
        df_ub["UB_First_Name"] = df_ub[ub_name_col].apply(lambda x: split_name(x)[0])
        df_ub["UB_Last_Name"]  = df_ub[ub_name_col].apply(lambda x: split_name(x)[1])
        ok_(f"Split '{ub_name_col}' -> UB_First_Name / UB_Last_Name")

    if ub_email_col:
        idx_pos = df_ub.columns.get_loc(ub_email_col)
        df_ub.insert(idx_pos+1, "UB_Email_Local",
                     df_ub[ub_email_col].apply(extract_local))
        df_ub.insert(idx_pos+2, "UB_Email_First",
                     df_ub["UB_Email_Local"].apply(lambda x: local_to_names(x)[0]))
        df_ub.insert(idx_pos+3, "UB_Email_Last",
                     df_ub["UB_Email_Local"].apply(lambda x: local_to_names(x)[1]))
        ok_(f"Added UB_Email_Local / UB_Email_First / UB_Email_Last after '{ub_email_col}'")

    if ub_sav_col:
        df_ub["UB_Saviynt_Local"] = df_ub[ub_sav_col].apply(extract_local)
        ok_(f"Added UB_Saviynt_Local from '{ub_sav_col}'")

    if ub_ad_col:
        df_ub["UB_AD_Local"] = df_ub[ub_ad_col].apply(extract_local)
        ok_(f"Added UB_AD_Local from '{ub_ad_col}'")

    st.update(s5, "OK", processed=len(df_ub))

    # ---- STEP 6: Build Userbase lookup indexes ------------------------------
    log("\n[STEP 6] Building Userbase lookup indexes ...")
    s6 = st.add("Build Userbase lookup indexes for fast matching")

    idx_email   = build_index(df_ub[ub_email_col], df_ub) if ub_email_col else {}
    idx_sav     = build_index(df_ub[ub_sav_col],   df_ub) if ub_sav_col   else {}
    idx_ad      = build_index(df_ub[ub_ad_col],    df_ub) if ub_ad_col    else {}
    idx_eml_loc = build_index(df_ub.get("UB_Email_Local",   pd.Series(dtype=str)), df_ub)
    idx_sav_loc = build_index(df_ub.get("UB_Saviynt_Local", pd.Series(dtype=str)), df_ub)
    idx_ad_loc  = build_index(df_ub.get("UB_AD_Local",      pd.Series(dtype=str)), df_ub)

    # Pre-build name index: "first||last" -> row dict  (O(1) Tier-3 lookup)
    idx_name = {}
    for _, ubrow in df_ub.iterrows():
        ub_f = norm(ubrow.get("UB_First_Name", ""))
        ub_l = norm(ubrow.get("UB_Last_Name",  ""))
        if ub_f and ub_l:
            key = f"{ub_f}||{ub_l}"
            if key not in idx_name:
                idx_name[key] = ubrow.to_dict()

    ok_(f"Email:{len(idx_email)}  Saviynt:{len(idx_sav)}  AD:{len(idx_ad)}")
    ok_(f"Eml-local:{len(idx_eml_loc)}  Sav-local:{len(idx_sav_loc)}  AD-local:{len(idx_ad_loc)}")
    ok_(f"Name index: {len(idx_name)} unique first+last combinations")
    st.update(s6, "OK", notes=f"email:{len(idx_email)} sav:{len(idx_sav)} "
                               f"ad:{len(idx_ad)} name:{len(idx_name)}")

    # ---- STEP 7: Multi-tier Userbase matching --------------------------------
    log("\n[STEP 7] Running multi-tier Userbase matching ...")
    s7 = st.add("Multi-tier Userbase matching (100% -> 75% -> 50%)")

    def get_ub_vals(row_dict):
        sav = row_dict.get(ub_sav_col, "") if ub_sav_col else ""
        ad  = row_dict.get(ub_ad_col,  "") if ub_ad_col  else ""
        return sav, ad

    results  = []
    cnt      = {100: 0, 75: 0, 50: 0, 0: 0}
    total    = len(df_gp)
    TICK     = max(1, total // 200)
    t0       = datetime.now()

    for i, (_, grow) in enumerate(df_gp.iterrows()):
        gp_email = norm(grow.get("email", ""))
        gp_local = norm(grow.get("_local", ""))
        gp_first = norm(grow.get("GP_First_Name", ""))
        gp_last  = norm(grow.get("GP_Last_Name",  ""))

        match_sav = match_ad = ""
        confidence = 0
        just = how = ""

        if   gp_email in idx_email:
            match_sav, match_ad = get_ub_vals(idx_email[gp_email])
            confidence, just, how = 100, "Email to Email (exact)", "email->email"
        elif gp_email in idx_sav:
            match_sav, match_ad = get_ub_vals(idx_sav[gp_email])
            confidence, just, how = 100, "Email matched Saviynt SSOUPN", "email->saviynt"
        elif gp_email in idx_ad:
            match_sav, match_ad = get_ub_vals(idx_ad[gp_email])
            confidence, just, how = 100, "Email matched AD SSOUPN", "email->ad"
        elif gp_local and gp_local in idx_eml_loc:
            match_sav, match_ad = get_ub_vals(idx_eml_loc[gp_local])
            confidence, just, how = 75, "Non-domain part matched UB_Email_Local", "local->email_local"
        elif gp_local and gp_local in idx_sav_loc:
            match_sav, match_ad = get_ub_vals(idx_sav_loc[gp_local])
            confidence, just, how = 75, "Non-domain part matched Saviynt local", "local->saviynt_local"
        elif gp_local and gp_local in idx_ad_loc:
            match_sav, match_ad = get_ub_vals(idx_ad_loc[gp_local])
            confidence, just, how = 75, "Non-domain part matched AD local", "local->ad_local"
        elif gp_first and gp_last:
            nk = f"{gp_first}||{gp_last}"
            if nk in idx_name:
                match_sav, match_ad = get_ub_vals(idx_name[nk])
                confidence, just, how = 50, "First and Last name matched", "name->name"

        cnt[confidence if confidence else 0] += 1
        results.append({
            "SSOUPN_Saviynt": match_sav,
            "SSOUPN_AD":      match_ad,
            "Confidence":     f"{confidence}%" if confidence else "No Match",
            "Justification":  just,
            "Match_Method":   how,
        })

        if i % TICK == 0 or i == total - 1:
            progress_bar(i, total, t0,
                         f"| 100%:{cnt[100]} 75%:{cnt[75]} 50%:{cnt[50]} none:{cnt[0]}")

    print()
    elapsed_ub = (datetime.now() - t0).total_seconds()
    df_match   = pd.DataFrame(results)
    df_out     = pd.concat([df_gp.reset_index(drop=True),
                             df_match.reset_index(drop=True)], axis=1)
    df_out.drop(columns=["_local"], inplace=True, errors="ignore")

    ok_(f"Done in {elapsed_ub:.1f}s -- 100%:{cnt[100]}  75%:{cnt[75]}  "
        f"50%:{cnt[50]}  none:{cnt[0]}")
    st.update(s7, "OK", processed=len(df_out), matched=cnt[100]+cnt[75]+cnt[50],
              notes=f"100%:{cnt[100]} 75%:{cnt[75]} 50%:{cnt[50]} "
                    f"none:{cnt[0]} ({elapsed_ub:.1f}s)")

    # ---- STEP 8: Load CSAM files --------------------------------------------
    log("\n[STEP 8] Loading CSAM log files ...")
    s8 = st.add("Load CSAM Submit & Clicked log files")

    def load_csam(filename, label):
        path = os.path.join(base_dir, filename)
        if not os.path.exists(path):
            err_(f"{label} not found: {path}")
            return None
        try:
            df = pd.read_excel(path, dtype=str)
            df.columns = [c.strip() for c in df.columns]
            ok_(f"Loaded {label}: {len(df)} rows | cols: {list(df.columns)}")
            return df
        except Exception as e:
            err_(f"Failed to load {label}: {e}")
            return None

    df_csam_sub = load_csam(CSAM_SUBMIT_FILENAME,  "CSAM Submit Logs")
    df_csam_clk = load_csam(CSAM_CLICKED_FILENAME, "CSAM Clicked Logs")

    if df_csam_sub is None and df_csam_clk is None:
        st.update(s8, "ERR", notes="Both CSAM files missing")
    elif df_csam_sub is None or df_csam_clk is None:
        missing = CSAM_SUBMIT_FILENAME if df_csam_sub is None else CSAM_CLICKED_FILENAME
        st.update(s8, "WARN", notes=f"Missing: {missing}")
    else:
        st.update(s8, "OK", notes="Both CSAM files loaded")

    # ---- STEP 9: Build CSAM indexes -----------------------------------------
    log("\n[STEP 9] Building CSAM lookup indexes ...")
    s9 = st.add("Build CSAM indexes (Email/UPN + local parts)")

    # Submit index: norm key -> {csam_email_upn, csam_username}
    csam_sub_idx = {}

    if df_csam_sub is not None:
        upn_col_s  = find_col(df_csam_sub, CSAM_COL_EMAIL_UPN, "CSAM Submit Email/UPN")
        user_col_s = find_col(df_csam_sub, CSAM_COL_USERNAME,  "CSAM Submit Username")

        for _, row in df_csam_sub.iterrows():
            upn_val  = str(row.get(upn_col_s,  "") if upn_col_s  else "").strip()
            user_val = str(row.get(user_col_s, "") if user_col_s else "").strip()
            rec = {"csam_email_upn": upn_val, "csam_username": user_val}
            # Index on full value AND local part for both email and username
            for candidate in [upn_val, extract_local(upn_val),
                               user_val, extract_local(user_val)]:
                k = norm(candidate)
                if k and k not in csam_sub_idx:
                    csam_sub_idx[k] = rec

        ok_(f"CSAM submit index: {len(csam_sub_idx)} keys")

    # Clicked index: norm key -> original Email/UPN value
    csam_clk_idx = {}

    if df_csam_clk is not None:
        upn_col_c = find_col(df_csam_clk, CSAM_COL_EMAIL_UPN, "CSAM Clicked Email/UPN")

        for _, row in df_csam_clk.iterrows():
            upn_val = str(row.get(upn_col_c, "") if upn_col_c else "").strip()
            for candidate in [upn_val, extract_local(upn_val)]:
                k = norm(candidate)
                if k and k not in csam_clk_idx:
                    csam_clk_idx[k] = upn_val

        ok_(f"CSAM clicked index: {len(csam_clk_idx)} keys")

    st.update(s9, "OK",
              notes=f"submit keys:{len(csam_sub_idx)}  clicked keys:{len(csam_clk_idx)}")

    # ---- STEP 10: Apply CSAM matching ---------------------------------------
    log("\n[STEP 10] Matching GoPhish rows against CSAM logs ...")
    s10 = st.add("Match GoPhish rows to CSAM Submit / Clicked logs")

    phished_status_col = []
    csam_email_upn_col = []
    csam_username_col  = []
    cnt_sub = cnt_clk = cnt_no_csam = 0

    t1    = datetime.now()
    TICK2 = max(1, len(df_out) // 100)

    for i, (_, row) in enumerate(df_out.iterrows()):
        # Build all candidate lookup keys for this GoPhish row
        gp_email  = norm(row.get("email",          ""))
        sav_full  = norm(row.get("SSOUPN_Saviynt", ""))
        ad_full   = norm(row.get("SSOUPN_AD",      ""))
        sav_local = extract_local(row.get("SSOUPN_Saviynt", ""))
        ad_local  = extract_local(row.get("SSOUPN_AD",      ""))
        gp_local  = extract_local(row.get("email", ""))

        candidates = [k for k in [gp_email, sav_full, ad_full,
                                   sav_local, ad_local, gp_local] if k]

        status_val = upn_val = uname_val = ""

        # Submit has priority over Clicked
        for k in candidates:
            if k in csam_sub_idx:
                rec        = csam_sub_idx[k]
                status_val = "Submitted Data"
                upn_val    = rec["csam_email_upn"]
                uname_val  = rec["csam_username"]
                cnt_sub   += 1
                break

        if not status_val:
            for k in candidates:
                if k in csam_clk_idx:
                    status_val = "Clicked Link"
                    upn_val    = csam_clk_idx[k]
                    cnt_clk   += 1
                    break

        if not status_val:
            cnt_no_csam += 1

        phished_status_col.append(status_val)
        csam_email_upn_col.append(upn_val)
        csam_username_col.append(uname_val)

        if i % TICK2 == 0 or i == len(df_out) - 1:
            progress_bar(i, len(df_out), t1,
                         f"| Submitted:{cnt_sub}  Clicked:{cnt_clk}  Unmatched:{cnt_no_csam}")

    print()
    elapsed_csam = (datetime.now() - t1).total_seconds()

    df_out["CSAM_Email_UPN"] = csam_email_upn_col
    df_out["CSAM_Username"]  = csam_username_col
    df_out["Phished_Status"] = phished_status_col

    ok_(f"Done in {elapsed_csam:.1f}s -- Submitted:{cnt_sub}  "
        f"Clicked:{cnt_clk}  No match:{cnt_no_csam}")
    st.update(s10, "OK", processed=len(df_out), matched=cnt_sub+cnt_clk,
              notes=f"Submitted:{cnt_sub} Clicked:{cnt_clk} "
                    f"NoMatch:{cnt_no_csam} ({elapsed_csam:.1f}s)")

    # ---- STEP 11: Write GoPhish Combined sheet ------------------------------
    log("\n[STEP 11] Writing GoPhish Combined sheet ...")
    s11 = st.add("Write GoPhish Combined sheet")

    # ---- Pre-write insights ----
    print()
    info_("--- GoPhish Combined Sheet Insights ---")
    print()

    # Campaign breakdown
    camp_counts = df_out["Campaign Name"].value_counts()
    info_(f"Campaigns: {len(camp_counts)}")
    for camp_name, camp_n in camp_counts.items():
        info_(f"  {camp_name}: {camp_n:,} rows")

    # Unique emails
    n_unique_emails = df_out["email"].nunique()
    info_(f"Total rows: {len(df_out):,}  |  Unique emails: {n_unique_emails:,}")
    print()

    # Confidence breakdown (overall)
    conf_series = df_out.get("Confidence", pd.Series(dtype=str))
    conf_vc = conf_series.value_counts()
    info_("Identity Match Confidence (overall):")
    for lbl in ["100%", "75%", "50%", "No Match"]:
        n = conf_vc.get(lbl, 0)
        pct = n / len(df_out) * 100 if len(df_out) else 0
        bar_w = int(pct / 100 * 30)
        bar = "#" * bar_w + "." * (30 - bar_w)
        info_(f"  {lbl:<10} [{bar}] {n:>6,}  ({pct:5.1f}%)")
    print()

    # Per-campaign confidence breakdown
    info_("Per-Campaign Identity Resolution:")
    for camp_name in camp_counts.index:
        camp_df = df_out[df_out["Campaign Name"] == camp_name]
        c_vc = camp_df["Confidence"].value_counts()
        c100 = c_vc.get("100%", 0)
        c75  = c_vc.get("75%", 0)
        c50  = c_vc.get("50%", 0)
        c0   = c_vc.get("No Match", 0)
        resolved = c100 + c75 + c50
        res_pct = resolved / len(camp_df) * 100 if len(camp_df) else 0
        info_(f"  {camp_name}")
        info_(f"    Resolved: {resolved:,}/{len(camp_df):,} ({res_pct:.1f}%)  "
              f"| 100%:{c100:,}  75%:{c75:,}  50%:{c50:,}  None:{c0:,}")
    print()

    # CSAM Phishing status breakdown
    phish_series = df_out.get("Phished_Status", pd.Series(dtype=str))
    info_("CSAM Phishing Status (overall):")
    n_submitted = (phish_series == "Submitted Data").sum()
    n_clicked   = (phish_series == "Clicked Link").sum()
    n_none      = ((phish_series == "") | phish_series.isna()).sum()
    for lbl, n, color_tag in [("Submitted Data", n_submitted, "!!"),
                               ("Clicked Link",   n_clicked,   "!"),
                               ("No CSAM match",  n_none,      "")]:
        pct = n / len(df_out) * 100 if len(df_out) else 0
        bar_w = int(pct / 100 * 30)
        bar = "#" * bar_w + "." * (30 - bar_w)
        info_(f"  {lbl:<16} [{bar}] {n:>6,}  ({pct:5.1f}%)")
    print()

    # Per-campaign phishing status
    info_("Per-Campaign Phishing Outcomes:")
    for camp_name in camp_counts.index:
        camp_df = df_out[df_out["Campaign Name"] == camp_name]
        ps = camp_df.get("Phished_Status", pd.Series(dtype=str))
        cs = ps.value_counts()
        c_sub = cs.get("Submitted Data", 0)
        c_clk = cs.get("Clicked Link", 0)
        c_no  = len(camp_df) - c_sub - c_clk
        risk_pct = (c_sub + c_clk) / len(camp_df) * 100 if len(camp_df) else 0
        info_(f"  {camp_name}")
        info_(f"    Submitted:{c_sub:,}  Clicked:{c_clk:,}  "
              f"Clean:{c_no:,}  | Risk rate: {risk_pct:.1f}%")
    print()

    # Match method distribution
    method_series = df_out.get("Match_Method", pd.Series(dtype=str))
    method_vc = method_series.value_counts()
    if len(method_vc):
        info_("Match Method Distribution:")
        for method, n in method_vc.items():
            if not method:
                method = "(no match)"
            pct = n / len(df_out) * 100 if len(df_out) else 0
            info_(f"  {method:<25} {n:>6,}  ({pct:5.1f}%)")
    print()

    info_("--- Writing sheet now (this may take a moment) ---")
    print()

    ws_gp = wb.create_sheet("GoPhish Combined")
    ws_gp.sheet_view.showGridLines = False

    GP_ORIG   = ["Campaign Name","campaign_id","email","time","message","details",
                 "GP_First_Name","GP_Last_Name"]
    UB_MATCH  = ["SSOUPN_Saviynt","SSOUPN_AD","Confidence","Justification","Match_Method"]
    CSAM_COLS = ["CSAM_Email_UPN","CSAM_Username","Phished_Status"]

    ALL_COLS = ([c for c in GP_ORIG   if c in df_out.columns] +
                [c for c in UB_MATCH  if c in df_out.columns] +
                [c for c in CSAM_COLS if c in df_out.columns])

    n_gp   = sum(1 for c in GP_ORIG   if c in df_out.columns)
    n_ub   = sum(1 for c in UB_MATCH  if c in df_out.columns)
    n_csam = sum(1 for c in CSAM_COLS if c in df_out.columns)

    # Row 1: section banners
    sections = [
        (1,                  n_gp,                "GoPhish Campaign Data",  C["gp_hdr"],  C["white"]),
        (n_gp+1,             n_gp+n_ub,           "Userbase Match",         C["nav_dark"],C["white"]),
        (n_gp+n_ub+1,        n_gp+n_ub+n_csam,   "CSAM Phishing Status",   "2D6A4F",     C["white"]),
    ]
    for start, end, label, bg, fg in sections:
        if n_csam == 0 and "CSAM" in label:
            continue
        if end >= start:
            ws_gp.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws_gp.cell(1, start)
            cell.value     = label
            cell.font      = mkfont(bold=True, color=fg, size=11)
            cell.fill      = mkfill(bg)
            cell.alignment = mkalign("center", "center")
    ws_gp.row_dimensions[1].height = 20

    # Row 2: column headers
    for ci, col_name in enumerate(ALL_COLS, 1):
        cell = ws_gp.cell(row=2, column=ci)
        cell.value     = col_name.replace("_", " ")
        cell.alignment = mkalign("center", "center", wrap=True)
        cell.border    = BORD
        if ci <= n_gp:
            cell.font = mkfont(bold=True, color=C["white"])
            cell.fill = mkfill(C["gp_hdr"])
        elif ci <= n_gp + n_ub:
            cell.font = mkfont(bold=True, color=C["white"])
            cell.fill = mkfill(C["nav_dark"])
        else:
            cell.font = mkfont(bold=True, color=C["white"])
            cell.fill = mkfill("2D6A4F")
    ws_gp.row_dimensions[2].height = 22
    ws_gp.freeze_panes = "A3"

    conf_bg = {"100%": C["map_100"], "75%": C["map_75"],
               "50%":  C["map_50"], "No Match": C["map_none"]}
    csam_bg = {"Submitted Data": C["csam_sub"],
               "Clicked Link":   C["csam_clk"]}

    total_rows = len(df_out)
    TICK3 = max(1, total_rows // 50)
    t_write = datetime.now()

    for ri, (_, row) in enumerate(df_out[ALL_COLS].iterrows(), 3):
        conf_val    = str(row.get("Confidence",     ""))
        csam_val    = str(row.get("Phished_Status", ""))
        alt_bg      = C["row_alt"] if ri % 2 == 0 else C["white"]
        match_bg    = conf_bg.get(conf_val, C["white"])
        csam_row_bg = csam_bg.get(csam_val, C["csam_none"] if csam_val == "" else C["white"])

        for ci, col_name in enumerate(ALL_COLS, 1):
            cell = ws_gp.cell(row=ri, column=ci)
            cell.value     = row[col_name]
            cell.border    = BORD
            cell.alignment = mkalign("left", "center")
            if ci <= n_gp:
                cell.font = mkfont(size=10, color="333333")
                cell.fill = mkfill(alt_bg)
            elif ci <= n_gp + n_ub:
                cell.font = mkfont(size=10)
                cell.fill = mkfill(match_bg)
            else:
                cell.font = mkfont(size=10)
                cell.fill = mkfill(csam_row_bg)

        row_idx = ri - 3
        if row_idx % TICK3 == 0 or row_idx == total_rows - 1:
            progress_bar(row_idx, total_rows, t_write, "| writing cells")

    print()
    elapsed_write = (datetime.now() - t_write).total_seconds()
    auto_width(ws_gp)
    ok_(f"Written {len(df_out):,} rows x {len(ALL_COLS)} cols "
        f"({len(ALL_COLS) * len(df_out):,} cells) to 'GoPhish Combined' "
        f"in {elapsed_write:.1f}s")
    st.update(s11, "OK", processed=len(df_out),
              notes=f"{len(df_out):,} rows, {len(ALL_COLS)} cols, "
                    f"{len(camp_counts)} campaigns ({elapsed_write:.1f}s)")

    # ---- STEP 12: Write Userbase Enriched sheet -----------------------------
    log("\n[STEP 12] Writing Userbase Enriched sheet ...")
    s12 = st.add("Write Userbase Enriched sheet")

    ws_ub = wb.create_sheet("Userbase Enriched")
    ws_ub.sheet_view.showGridLines = False
    ws_ub.freeze_panes = "A2"
    ub_cols = list(df_ub.columns)
    write_header_row(ws_ub, [c.replace("_", " ") for c in ub_cols])

    for ri, (_, row) in enumerate(df_ub.iterrows(), 2):
        alt = ri % 2 == 0
        for ci, col in enumerate(ub_cols, 1):
            cell = ws_ub.cell(row=ri, column=ci)
            cell.value     = row[col]
            cell.font      = mkfont(size=10)
            cell.fill      = mkfill(C["row_alt"] if alt else C["white"])
            cell.border    = BORD
            cell.alignment = mkalign("left", "center")

    auto_width(ws_ub)
    ok_(f"Written {len(df_ub)} rows to 'Userbase Enriched'")
    st.update(s12, "OK", processed=len(df_ub))

    # ---- STEP 13: Save ------------------------------------------------------
    log("\n[STEP 13] Saving workbook ...")
    s13 = st.add("Save output workbook")

    if "_tmp" in wb.sheetnames:
        del wb["_tmp"]
    order = ["Status", "GoPhish Combined", "Userbase Enriched"]
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    wb.save(output_path)
    ok_(f"Saved -> {output_path}")
    st.update(s13, "OK", notes=output_path)
    wb.save(output_path)

    # ---- Final summary -------------------------------------------------------
    print("\n" + "=" * 68)
    log("COMPLETE -- Summary")
    info_(f"GoPhish rows      : {len(df_gp)}")
    info_(f"Unique emails     : {df_gp['email'].nunique()}")
    info_(f"UB 100% match     : {cnt[100]}")
    info_(f"UB  75% match     : {cnt[75]}")
    info_(f"UB  50% match     : {cnt[50]}")
    info_(f"UB  no match      : {cnt[0]}")
    info_(f"CSAM Submitted    : {cnt_sub}")
    info_(f"CSAM Clicked      : {cnt_clk}")
    info_(f"CSAM no match     : {cnt_no_csam}")
    info_(f"Output            : {output_path}")
    print("=" * 68)


if __name__ == "__main__":
    main()
