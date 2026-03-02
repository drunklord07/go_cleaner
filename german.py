#!/usr/bin/env python3
"""
GoPhish Campaign Events Processor
Combines, categorizes, and visualizes phishing campaign data.
"""

import os
import sys
import glob
import json
import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

# ─── Color Palette ────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":      "1F3864",   # dark navy
    "header_fg":      "FFFFFF",
    "submitted_dark":  "FF9999",  # red
    "submitted_light": "FFD6D6",
    "clicked_dark":    "FFD580",  # amber
    "clicked_light":   "FFF3CD",
    "opened_dark":     "90EE90",  # green
    "opened_light":    "D6F5D6",
    "sent_dark":       "C0C0C0",  # grey
    "sent_light":      "EFEFEF",
    "status_done":     "C6EFCE",
    "status_error":    "FFC7CE",
    "status_skip":     "FFEB9C",
    "status_header":   "2E4057",
    "trail_alt1":      "F7F7F7",
}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# GoPhish message values (case-insensitive match)
MSG_SUBMITTED   = "submitted data"
MSG_CLICKED     = "clicked link"
MSG_OPENED      = "email opened"
MSG_SENT        = "email sent"
MSG_CREATED     = "campaign created"

MESSAGE_ORDER = [MSG_SUBMITTED, MSG_CLICKED, MSG_OPENED, MSG_SENT, MSG_CREATED]

# ─── Helpers ──────────────────────────────────────────────────────────────────

def print_step(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def style_header_row(ws, row_idx, col_count, bg=COLORS["header_bg"], fg=COLORS["header_fg"]):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = make_fill(bg)
        cell.font = Font(bold=True, color=fg, name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 20

def auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def normalize_msg(val):
    if pd.isna(val):
        return ""
    return str(val).strip().lower()

def get_email_rank(email_df):
    """Return highest-priority message for an email."""
    msgs = set(email_df["message_norm"].tolist())
    for m in MESSAGE_ORDER:
        if m in msgs:
            return m
    return MSG_SENT

# ─── Status Sheet ─────────────────────────────────────────────────────────────

class StatusSheet:
    def __init__(self, wb):
        self.ws = wb.create_sheet("Status", 0)
        self._setup()
        self.steps = []

    def _setup(self):
        ws = self.ws
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 38

        title_cell = ws["B1"]
        title_cell.value = "GoPhish Processor – Run Status"
        title_cell.font = Font(bold=True, size=14, color=COLORS["header_fg"], name="Arial")
        ws["B1"].fill = make_fill(COLORS["status_header"])
        ws.merge_cells("B1:E1")
        ws.row_dimensions[1].height = 28

        ws["B2"].value = f"Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["B2"].font = Font(italic=True, size=9, name="Arial")

        headers = ["Step", "Description", "Status", "Count", "Notes"]
        for i, h in enumerate(headers, 1):
            col = get_column_letter(i + 1)  # B onwards
            ws[f"{col}4"].value = h
        style_header_row(ws, 4, 5, bg=COLORS["status_header"])
        # shift: actual cols B-F = 2-6
        for i, h in enumerate(headers, 2):
            col = get_column_letter(i)
            ws[f"{col}4"].value = h
        ws.freeze_panes = "B5"
        self.next_row = 5

    def log(self, description, status="⏳ Running", count=None, notes=""):
        ws = self.ws
        row = self.next_row
        ws.cell(row=row, column=2).value = len(self.steps) + 1
        ws.cell(row=row, column=3).value = description
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = count if count is not None else ""
        ws.cell(row=row, column=6).value = notes
        for col in range(2, 7):
            ws.cell(row=row, column=col).font = Font(name="Arial", size=10)
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        self.steps.append({"row": row, "desc": description})
        self.next_row += 1
        return row  # so caller can update later

    def update(self, row, status, count=None, notes=None):
        ws = self.ws
        ws.cell(row=row, column=4).value = status
        if count is not None:
            ws.cell(row=row, column=5).value = count
        if notes is not None:
            ws.cell(row=row, column=6).value = notes

        color = COLORS["status_done"] if "✅" in status else (
            COLORS["status_error"] if "❌" in status else COLORS["status_skip"]
        )
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = make_fill(color)

# ─── Write DataFrame to Sheet ─────────────────────────────────────────────────

def df_to_sheet(ws, df, header_cols=None):
    """Write df to ws starting at row 1. Returns row count written."""
    cols = header_cols or list(df.columns)
    for c_idx, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=c_idx).value = col_name
    style_header_row(ws, 1, len(cols))
    ws.freeze_panes = f"A2"

    for r_idx, (_, row_data) in enumerate(df[cols].iterrows(), 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    auto_width(ws)
    return len(df)

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir  = os.path.join(script_dir, "input")
    output_path = os.path.join(script_dir, "gophish_report.xlsx")

    print_step("Initialising workbook and Status sheet …")
    wb = Workbook()
    # Remove default sheet later after we build everything
    default_ws = wb.active
    default_ws.title = "_tmp"

    status = StatusSheet(wb)
    print_step("Status sheet created.")

    # ── Step 1: Combine input files ─────────────────────────────────────────
    s1 = status.log("Scan input/ folder for CSV files")
    excel_files = sorted(glob.glob(os.path.join(input_dir, "*.csv")))
    if not excel_files:
        status.update(s1, "❌ Error", 0, f"No CSV files found in {input_dir}")
        print_step(f"ERROR: No CSV files found in {input_dir}. Aborting.")
        wb.save(output_path)
        sys.exit(1)
    status.update(s1, "✅ Done", len(excel_files), f"Found {len(excel_files)} CSV file(s)")
    print_step(f"Found {len(excel_files)} CSV input file(s).")

    s2 = status.log("Combine all input files into one DataFrame")
    try:
        frames = []
        for fp in excel_files:
            print_step(f"  Reading: {os.path.basename(fp)}")
            tmp = pd.read_csv(fp, dtype=str, encoding="utf-8-sig")
            tmp.columns = [c.strip() for c in tmp.columns]
            frames.append(tmp)
        df_all = pd.concat(frames, ignore_index=True)
        # Normalize expected columns
        col_map = {}
        for c in df_all.columns:
            cl = c.lower().replace(" ", "_")
            if "campaign" in cl:  col_map[c] = "campaign_id"
            elif "email" in cl:   col_map[c] = "email"
            elif "time" in cl:    col_map[c] = "time"
            elif "message" in cl: col_map[c] = "message"
            elif "detail" in cl:  col_map[c] = "details"
        df_all.rename(columns=col_map, inplace=True)
        EXPECTED = ["campaign_id", "email", "time", "message", "details"]
        for col in EXPECTED:
            if col not in df_all.columns:
                df_all[col] = ""
        df_all = df_all[EXPECTED]
        df_all["message_norm"] = df_all["message"].apply(normalize_msg)
        status.update(s2, "✅ Done", len(df_all), f"Total rows: {len(df_all)}")
        print_step(f"Combined DataFrame: {len(df_all)} rows.")
    except Exception as e:
        status.update(s2, "❌ Error", notes=str(e))
        print_step(f"ERROR combining files: {e}")
        wb.save(output_path)
        sys.exit(1)

    # ── Step 2: Parse details JSON ──────────────────────────────────────────
    s3 = status.log("Parse 'details' JSON column")
    def parse_details(raw):
        if pd.isna(raw) or str(raw).strip() in ("", "nan"):
            return raw
        try:
            d = json.loads(str(raw))
            # GoPhish payload format: {"payload": {"key": ["val"]}, ...}
            payload = d.get("payload", d)
            parts = []
            for k, v in payload.items():
                if k.lower() in ("rid", "__orig_pwd"):
                    continue
                val = v[0] if isinstance(v, list) and v else v
                parts.append(f"{k}: {val}")
            return " | ".join(parts) if parts else raw
        except Exception:
            return raw
    df_all["details_parsed"] = df_all["details"].apply(parse_details)
    status.update(s3, "✅ Done", notes="details_parsed column added")
    print_step("Details JSON parsed.")

    # ── Step 3: Build per-email category ────────────────────────────────────
    s4 = status.log("Categorise each email by highest event reached")
    email_rank = {}
    for email, grp in df_all.groupby("email"):
        email_rank[email] = get_email_rank(grp)
    submitted_emails = {e for e, r in email_rank.items() if r == MSG_SUBMITTED}
    clicked_emails   = {e for e, r in email_rank.items() if r == MSG_CLICKED}
    opened_emails    = {e for e, r in email_rank.items() if r == MSG_OPENED}
    sent_emails      = {e for e, r in email_rank.items() if r == MSG_SENT or r == MSG_CREATED}
    status.update(s4, "✅ Done", notes=(
        f"Submitted:{len(submitted_emails)} Clicked:{len(clicked_emails)} "
        f"Opened:{len(opened_emails)} Sent/other:{len(sent_emails)}"
    ))
    print_step(f"Categories – Submitted:{len(submitted_emails)} Clicked:{len(clicked_emails)} "
               f"Opened:{len(opened_emails)} Sent:{len(sent_emails)}")

    # ── Step 4: Build Trails sheet ─────────────────────────────────────────
    s5 = status.log("Build Trails sheet (grouped journeys)")
    print_step("Building Trails sheet …")
    ws_trails = wb.create_sheet("Trails")
    ws_trails.sheet_view.showGridLines = False

    TRAIL_COLS = ["campaign_id", "email", "time", "message", "details_parsed"]
    TRAIL_DISPLAY = ["Campaign ID", "Email", "Time", "Message", "Details"]
    for c_idx, h in enumerate(TRAIL_DISPLAY, 1):
        ws_trails.cell(row=1, column=c_idx).value = h
    style_header_row(ws_trails, 1, len(TRAIL_COLS))
    ws_trails.freeze_panes = "A2"

    # Ordered group list: submitted → clicked → opened → sent
    def emails_for_rank(rank_set):
        return sorted(rank_set)

    group_order = [
        (submitted_emails, COLORS["submitted_dark"],  COLORS["submitted_light"]),
        (clicked_emails,   COLORS["clicked_dark"],    COLORS["clicked_light"]),
        (opened_emails,    COLORS["opened_dark"],     COLORS["opened_light"]),
        (sent_emails,      COLORS["sent_dark"],       COLORS["sent_light"]),
    ]

    # email → first row in Trails (for cross-sheet linking)
    email_trail_row = {}
    current_row = 2
    alt = False  # alternate shade within same group

    for (email_set, dark_color, light_color) in group_order:
        for email in emails_for_rank(email_set):
            grp = df_all[df_all["email"] == email].copy()
            # Sort by message priority (sent → opened → clicked → submitted)
            priority = {MSG_CREATED: 0, MSG_SENT: 1, MSG_OPENED: 2,
                        MSG_CLICKED: 3, MSG_SUBMITTED: 4}
            grp["_pri"] = grp["message_norm"].map(lambda m: priority.get(m, 5))
            grp.sort_values("_pri", inplace=True)
            grp.drop(columns="_pri", inplace=True)

            email_trail_row[email] = current_row
            fill = make_fill(dark_color if not alt else light_color)

            for _, row_data in grp.iterrows():
                for c_idx, col in enumerate(TRAIL_COLS, 1):
                    cell = ws_trails.cell(row=current_row, column=c_idx)
                    cell.value = row_data.get(col, "")
                    cell.fill = fill
                    cell.font = Font(name="Arial", size=10)
                    cell.border = BORDER
                    cell.alignment = Alignment(vertical="center")
                current_row += 1
            alt = not alt

    auto_width(ws_trails)
    status.update(s5, "✅ Done", current_row - 2, f"{current_row-2} trail rows written")
    print_step(f"Trails sheet done – {current_row-2} rows.")

    # ── Step 5: Build category sheets with hyperlink to Trails ────────────
    DISPLAY_COLS = ["campaign_id", "email", "time", "message", "details_parsed"]
    DISPLAY_HDRS = ["Campaign ID", "Email", "Time", "Message", "Details"]

    def build_category_sheet(name, email_set, bg_color):
        s = status.log(f"Build '{name}' sheet")
        print_step(f"Building '{name}' sheet …")
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False

        hdrs = DISPLAY_HDRS + ["Trail Link"]
        for c_idx, h in enumerate(hdrs, 1):
            ws.cell(row=1, column=c_idx).value = h
        style_header_row(ws, 1, len(hdrs))
        ws.freeze_panes = "A2"

        r = 2
        for email in sorted(email_set):
            grp = df_all[df_all["email"] == email]
            # One row per email (highest message row)
            row_data = grp[grp["message_norm"] == email_rank[email]].iloc[0]
            for c_idx, col in enumerate(DISPLAY_COLS, 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.value = row_data.get(col, "")
                cell.font = Font(name="Arial", size=10)
                cell.fill = make_fill(bg_color)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")

            # Hyperlink to Trails
            trail_row = email_trail_row.get(email, 1)
            link_cell = ws.cell(row=r, column=len(DISPLAY_COLS) + 1)
            link_cell.value = "View Journey"
            link_cell.hyperlink = f"#Trails!A{trail_row}"
            link_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            link_cell.fill = make_fill(bg_color)
            link_cell.border = BORDER
            r += 1

        auto_width(ws)
        status.update(s, "✅ Done", r - 2)
        print_step(f"'{name}' sheet done – {r-2} rows.")

    build_category_sheet("Submitted Data", submitted_emails, COLORS["submitted_light"])
    build_category_sheet("Clicked Link",   clicked_emails,  COLORS["clicked_light"])
    build_category_sheet("Email Opened",   opened_emails,   COLORS["opened_light"])
    build_category_sheet("Email Sent",     sent_emails,     COLORS["sent_light"])

    # ── Step 6: Main (Combined) sheet ───────────────────────────────────────
    s6 = status.log("Build Main (combined raw) sheet")
    print_step("Building Main sheet …")
    ws_main = wb.create_sheet("Main")
    ws_main.sheet_view.showGridLines = False
    df_display = df_all[DISPLAY_COLS].copy()
    df_display.columns = DISPLAY_HDRS
    written = df_to_sheet(ws_main, df_display)
    status.update(s6, "✅ Done", written)
    print_step(f"Main sheet done – {written} rows.")

    # ── Step 7: Summary stats in Status ─────────────────────────────────────
    s7 = status.log("Write summary statistics")
    total_emails = len(email_rank)
    status.update(s7, "✅ Done", total_emails, (
        f"Total unique emails: {total_emails} | "
        f"Submitted: {len(submitted_emails)} ({100*len(submitted_emails)//max(total_emails,1)}%) | "
        f"Clicked: {len(clicked_emails)} | Opened: {len(opened_emails)} | Sent: {len(sent_emails)}"
    ))

    # ── Finalise ────────────────────────────────────────────────────────────
    # Remove temp sheet
    if "_tmp" in wb.sheetnames:
        del wb["_tmp"]

    # Reorder sheets: Status first, then Trails, then categories, then Main
    desired_order = ["Status", "Trails", "Submitted Data", "Clicked Link",
                     "Email Opened", "Email Sent", "Main"]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    s8 = status.log("Save workbook")
    print_step(f"Saving to {output_path} …")
    wb.save(output_path)
    status_final_row = status.log("COMPLETE")
    status.update(s8, "✅ Done", notes=output_path)
    wb.save(output_path)
    print_step(f"✅ All done! Output: {output_path}")

if __name__ == "__main__":
    main()
